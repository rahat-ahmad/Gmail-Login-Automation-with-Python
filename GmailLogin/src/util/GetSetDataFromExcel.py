from os import path

from openpyxl import load_workbook


class GetSetDataFromExcel():

    def writingTestResult(self,cell, status):
        fileDir = path.abspath(path.join(__file__, "../../../resources/GmailLoginTestCase.xlsx"))
        workbook = load_workbook(filename=fileDir)

        # open workbook
        sheet = workbook.active

        # modify the desired cell
        sheet[cell] = status

        # save the file
        workbook.save(filename=fileDir)
        workbook.close()

    def readTestData(self,row, column):

        fileDir = path.abspath(path.join(__file__, "../../../resources/TestData.xlsx"))
        workbook = load_workbook(filename=fileDir)

        # open workbook
        sheet = workbook.active
        print(sheet.cell(row = row, column = column).value)
        return sheet.cell(row = row, column = column).value
        # modify the desired cell
